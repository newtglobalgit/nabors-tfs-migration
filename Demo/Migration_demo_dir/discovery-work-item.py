import winrm, threading, time, sys
import pandas as pd
import openpyxl
from credentials import cred, server, server_urls, projects

url = server.get('host')
tfs_url = server_urls.get('http_url')+'DefaultCollection'
project_name = projects.get('project5')
username = cred.get('USER')
password = cred.get('PASSWORD')
workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet['A1'] = 'Field'
worksheet['B1'] = 'Value'

session = winrm.Session(url, auth=(username, password), transport='ntlm')
script = """$url = '%s'
$projectName = '%s'
$query = @"
SELECT [System.Id], [System.WorkItemType], [System.State]
FROM WorkItems 
WHERE [System.TeamProject] = '$projectName'
"@
TFPT.EXE query /collection:$url /wiql:$query /include:data | Out-String -stream
""" % (tfs_url, project_name)

output = session.run_ps(script)

results = output.std_out.decode('utf-8').strip().split('\n')
print(results)
work_items = []
for result in results:
    fields = result.split('\t')
    work_item = {
        'id': fields[0],
        'type': fields[1],
        'state': fields[2]
    }
    work_items.append(work_item)

total_count = len(work_items)
type_counts = {}
closed_count = 0
open_count = 0

for work_item in work_items:
    if work_item['type'] in type_counts:
        type_counts[work_item['type']] += 1
    else:
        type_counts[work_item['type']] = 1

    if work_item['state'] == 'Closed':
        closed_count += 1
    else:
        open_count += 1

workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet['A1'] = 'WorkItem Count'
worksheet['B1'] = 'Value'
worksheet['A2'] = 'Total Count'
worksheet['B2'] = total_count
row_num = 3
for work_item_type, count in type_counts.items():
    worksheet.cell(row=row_num, column=1).value = f'{work_item_type} Count'
    worksheet.cell(row=row_num, column=2).value = count
    row_num += 1
worksheet.cell(row=row_num, column=1).value = 'Closed Count'
worksheet.cell(row=row_num, column=2).value = closed_count
row_num += 1
worksheet.cell(row=row_num, column=1).value = 'Open Count'
worksheet.cell(row=row_num, column=2).value = open_count
workbook.save('Discovery Report - WorkItems.xlsx')